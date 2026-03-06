import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
from typing import Optional, List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.database import get_db
from app.machines import calculate_validation
import config


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHEET_MAP = [
    ('small_machines',                    'Male stroje-zariadenia',    True),
    ('big_machines',                      'Velke stroje-zariadenia',   True),
    ('temporary_archived_small_machines', 'Male-docasne_vyradene',     False),
    ('archived_small_machines',           'Male-vyradene',             False),
    ('temporary_archived_big_machines',   'Velke-docasne-vyradene',    False),
    ('archived_big_machines',             'Velke-vyradene',            False),
]

# Sheets whose validation column should be colour-coded
ACTIVE_SHEETS = {'Male stroje-zariadenia', 'Velke stroje-zariadenia'}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _load_and_rename(table: str, add_validation: bool) -> pd.DataFrame:
    """Read a table into a DataFrame, optionally computing the validation column."""
    df = pd.read_sql_query(f'SELECT * FROM {table}', get_db())
    if add_validation:
        df['validation'] = df.apply(calculate_validation, axis=1)
    return df.rename(columns=config.COLUMN_RENAME_MAP)


def _write_excel(file_name: str, sheets: List[Tuple[str, str, bool]]) -> None:
    """Write all sheets to an Excel file and apply colour-coding."""
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    for table, sheet_name, add_validation in sheets:
        _load_and_rename(table, add_validation).to_excel(writer, sheet_name=sheet_name, index=False)
    writer._save()

    _apply_validation_colours(file_name)
    print(f'Excel file saved: {file_name}')


def _apply_validation_colours(file_name: str) -> None:
    """Colour-code the validation column in active sheets (red = overdue, yellow = due soon)."""
    workbook = load_workbook(file_name)
    today = datetime.now().date()

    for sheet_name in ACTIVE_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=13, max_col=13):
            for cell in row:
                val = cell.value
                if isinstance(val, datetime):
                    val = val.date()
                if val is None:
                    continue
                if val < today:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif (val - today).days <= 30:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    workbook.save(file_name)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_excel(output_path: Optional[str] = None) -> str:
    """Export all machine tables to an Excel file and return the file path."""
    day_name = datetime.now().strftime('%A')
    file_name = output_path or (
        f'{config.EXCEL_OUTPUT_DIR}\\Evidencia_Strojov_{day_name}.xlsx'
    )
    _write_excel(file_name, SHEET_MAP)
    return file_name


def send_weekly_email() -> None:
    """Generate a report of overdue machines and e-mail it to the configured recipients."""
    db = get_db()
    today = datetime.now()
    attachment_path = 'Evidencia_po_platnosti_revizie.xlsx'

    # Build overdue DataFrames
    small_df = pd.read_sql_query('SELECT * FROM small_machines', db)
    small_df['validation'] = small_df.apply(calculate_validation, axis=1)
    big_df = pd.read_sql_query('SELECT * FROM big_machines', db)
    big_df['validation'] = big_df.apply(calculate_validation, axis=1)

    overdue_small = small_df[small_df['validation'] < today].rename(columns=config.COLUMN_RENAME_MAP)
    overdue_big = big_df[big_df['validation'] < today].rename(columns=config.COLUMN_RENAME_MAP)

    with pd.ExcelWriter(attachment_path, engine='xlsxwriter') as writer:
        overdue_small.to_excel(writer, sheet_name='Male stroje-zariadenia', index=False)
        overdue_big.to_excel(writer, sheet_name='Velke stroje-zariadenia', index=False)

    _send_email(attachment_path)


def _send_email(attachment_path: str) -> None:
    """Send the weekly report email with the given file attached."""
    msg = MIMEMultipart()
    msg['From'] = config.EMAIL_SENDER
    msg['To'] = ', '.join(config.EMAIL_RECEIVERS)
    msg['Subject'] = 'Týždenný report'
    msg.attach(MIMEText(
        'V prílohe sa nachádza súbor so zariadeniami/strojmi po uplynutí platnosti revízie'
        '\n\n\n\nTáto správa je generovaná automaticky',
        'plain',
    ))

    with open(attachment_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={attachment_path}')
    msg.attach(part)

    with smtplib.SMTP(config.EMAIL_SMTP_SERVER, config.EMAIL_PORT) as server:
        server.starttls()
        server.login(config.EMAIL_USERNAME, config.EMAIL_PASSWORD)
        server.sendmail(config.EMAIL_SENDER, config.EMAIL_RECEIVERS, msg.as_string())

    print('Email sent successfully.')